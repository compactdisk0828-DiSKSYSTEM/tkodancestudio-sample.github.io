#!/usr/bin/env python3
"""
TKO DANCE STUDIO - Time Table Excel & Image Generator
Generates styled Excel timetables for Yatsuka and Kitakoshigaya studios,
then converts them to PNG images.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
IMAGES_DIR = os.path.join(OUTPUT_DIR, "images")
EXCEL_DIR = os.path.join(OUTPUT_DIR, "data")

os.makedirs(IMAGES_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

# ─── Color Constants ────────────────────────────────────────
WHITE = "FFFFFF"
BLACK = "111111"
RED = "FF0000"
NAVY = "000080"
CYAN = "00D4FF"
YELLOW = "FFE033"
LIGHT_GRAY = "F5F5F5"
DARK_BG = "1A1A1A"
HEADER_BG = "222222"

# ─── Schedule Data ──────────────────────────────────────────
yatsuka_schedule = {
    "MON": [
        {"time": "18:00-19:00", "name": "キッズ girl's HIP HOP\n入門・初級", "instructor": "Emi", "type": "kids"},
        {"time": "19:30-21:00", "name": "Waack", "instructor": "ちーこ", "type": "open"},
    ],
    "TUE": [
        {"time": "18:15-19:15", "name": "キッズ HIP HOP\n入門・初級", "instructor": "MIYUU", "type": "kids"},
        {"time": "19:30-21:00", "name": "HIP HOP 初中級", "instructor": "MIYUU", "type": "open"},
    ],
    "WED": [
        {"time": "18:15-19:15", "name": "キッズ JAZZ HIP HOP\n入門・初級", "instructor": "MOMOKA", "type": "kids"},
        {"time": "20:00-21:30", "name": "JAZZ HIP HOP\n初中級", "instructor": "MOMOKA", "type": "open"},
    ],
    "THU": [
        {"time": "18:15-19:45", "name": "JAZZ", "instructor": "イクミ", "type": "open"},
        {"time": "20:00-21:30", "name": "K-POP", "instructor": "Kurumi", "type": "open"},
    ],
    "FRI": [
        {"time": "17:20-18:10", "name": "リトル HIP HOP\n(年中から)", "instructor": "TKO", "type": "little"},
        {"time": "18:15-19:45", "name": "キッズ HIP HOP\n初級", "instructor": "TKO", "type": "kids"},
        {"time": "20:00-21:30", "name": "house", "instructor": "RUKA", "type": "open"},
    ],
    "SAT": [
        {"time": "09:45-10:45", "name": "キッズダンス\nスタートクラス", "instructor": "key", "type": "kids"},
        {"time": "11:00-12:30", "name": "HOUSE 初級", "instructor": "key", "type": "open"},
        {"time": "13:00-13:45", "name": "キッズ HIP HOP\n入門", "instructor": "Yura", "type": "little"},
        {"time": "14:00-15:00", "name": "HIP HOP\n初中級", "instructor": "Yura", "type": "kids"},
        {"time": "18:45-20:15", "name": "Jr. HIP HOP\n中級", "instructor": "TAKAO", "type": "open"},
        {"time": "20:30-22:00", "name": "HIP HOP 中級", "instructor": "TAKAO", "type": "open"},
    ],
    "SUN": [
        {"time": "18:00-19:30", "name": "girl's HIP HOP\n(承認制)", "instructor": "AYAKA", "type": "open"},
    ],
}

kitakoshigaya_schedule = {
    "MON": [],
    "TUE": [],
    "WED": [],
    "THU": [
        {"time": "18:00-19:00", "name": "HIP HOP", "instructor": "TKO", "type": "kids"},
    ],
    "FRI": [],
    "SAT": [],
    "SUN": [],
}


def get_type_fill(cls_type):
    """Return (bg_color, text_color) based on class type."""
    if cls_type == "little":
        return YELLOW, BLACK
    elif cls_type == "kids":
        return CYAN, BLACK
    else:  # open/adult
        return WHITE, RED


def create_timetable_excel(schedule, studio_name, filename, studio_color=RED, note_text=None, is_coming_soon=False):
    """Create a styled Excel timetable."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TIME TABLE"

    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    max_classes = max(len(schedule.get(d, [])) for d in days) if not is_coming_soon else 1
    if max_classes == 0:
        max_classes = 1

    # Each class uses 3 rows: time, name, instructor
    rows_per_class = 3
    total_class_rows = max_classes * rows_per_class

    # ─── Column widths ──────────────────────────────────
    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # ─── Styling constants ──────────────────────────────
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    thick_bottom = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='medium', color='999999'),
    )

    # ─── Row 1: Studio Title ────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=1, value=studio_name)
    title_cell.font = Font(name='Arial', bold=True, size=14, color=studio_color)
    title_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # ─── Row 2: Day headers ─────────────────────────────
    weekend_days = {"SAT": NAVY, "SUN": RED}
    for col_idx, day in enumerate(days, 1):
        cell = ws.cell(row=2, column=col_idx, value=day)
        if day in weekend_days:
            bg = weekend_days[day]
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            cell.font = Font(name='Arial', bold=True, size=11, color=WHITE)
        else:
            cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
            cell.font = Font(name='Arial', bold=True, size=11, color=BLACK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # ─── Class rows ─────────────────────────────────────
    for col_idx, day in enumerate(days, 1):
        classes = schedule.get(day, [])
        for cls_idx in range(max_classes):
            base_row = 3 + cls_idx * rows_per_class

            if cls_idx < len(classes):
                cls = classes[cls_idx]
                bg_color, text_color = get_type_fill(cls["type"])
                fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

                # Time row
                time_cell = ws.cell(row=base_row, column=col_idx, value=cls["time"])
                time_cell.font = Font(name='Arial', size=8, color=text_color if text_color != RED else BLACK)
                time_cell.fill = fill
                time_cell.alignment = Alignment(horizontal='center', vertical='center')
                time_cell.border = thin_border

                # Name row
                name_cell = ws.cell(row=base_row + 1, column=col_idx, value=cls["name"])
                name_cell.font = Font(name='Arial', bold=True, size=9, color=text_color)
                name_cell.fill = fill
                name_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                name_cell.border = thin_border

                # Instructor row
                instr_cell = ws.cell(row=base_row + 2, column=col_idx, value=cls["instructor"])
                instr_cell.font = Font(name='Arial', bold=True, size=9, color=text_color if text_color != RED else BLACK)
                instr_cell.fill = fill
                instr_cell.alignment = Alignment(horizontal='center', vertical='bottom')
                instr_cell.border = thick_bottom
            else:
                # Empty slots
                for r_offset in range(rows_per_class):
                    empty_cell = ws.cell(row=base_row + r_offset, column=col_idx)
                    empty_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')
                    empty_cell.border = thin_border

    # Set row heights for class rows
    for cls_idx in range(max_classes):
        base_row = 3 + cls_idx * rows_per_class
        ws.row_dimensions[base_row].height = 20       # time
        ws.row_dimensions[base_row + 1].height = 40   # name
        ws.row_dimensions[base_row + 2].height = 22   # instructor

    # ─── Note row (for Yatsuka Sunday) ──────────────────
    note_row = 3 + max_classes * rows_per_class
    if note_text:
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
        note_cell = ws.cell(row=note_row, column=1, value=note_text)
        note_cell.font = Font(name='Arial', size=8, color='888888')
        note_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')
        note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[note_row].height = 35

    # ─── Legend row ─────────────────────────────────────
    legend_row = note_row + 1 if note_text else note_row
    legends = [
        ("■ Little (KIDS)", YELLOW, BLACK),
        ("■ KIDS (Under 12)", CYAN, BLACK),
        ("■ Open / Adult", WHITE, BLACK),
    ]
    for i, (label, bg, fg) in enumerate(legends):
        col = 1 + i * 2
        if col <= 7:
            legend_cell = ws.cell(row=legend_row, column=col, value=label)
            legend_cell.font = Font(name='Arial', size=8, color=fg)
            legend_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')
            legend_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[legend_row].height = 25

    # Fill remaining legend cells with dark bg
    for col_idx in range(1, 8):
        cell = ws.cell(row=legend_row, column=col_idx)
        if cell.value is None:
            cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')

    filepath = os.path.join(EXCEL_DIR, filename)
    wb.save(filepath)
    print(f"✅ Created Excel: {filepath}")
    return filepath


def excel_to_image(excel_path, output_image_path):
    """
    Render an Excel timetable to a PNG image using Pillow.
    This reads the openpyxl workbook and draws cells manually.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # ─── Layout parameters ──────────────────────────────
    col_width_px = 160
    num_cols = 7
    padding = 20
    img_width = num_cols * col_width_px + padding * 2

    # Calculate row heights
    row_heights = {}
    for row_idx in range(1, ws.max_row + 1):
        h = ws.row_dimensions[row_idx].height
        row_heights[row_idx] = int(h * 1.8) if h else 30

    total_height = sum(row_heights.get(r, 30) for r in range(1, ws.max_row + 1)) + padding * 2

    img = Image.new('RGB', (img_width, total_height), color=(26, 26, 26))
    draw = ImageDraw.Draw(img)

    # Try to load fonts
    try:
        font_bold_large = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc", 22)
        font_bold = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc", 14)
        font_bold_small = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc", 12)
        font_small = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc", 10)
        font_tiny = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc", 9)
    except (OSError, IOError):
        try:
            font_bold_large = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Bold.ttf", 22)
            font_bold = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Bold.ttf", 14)
            font_bold_small = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Bold.ttf", 12)
            font_small = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 10)
            font_tiny = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 9)
        except (OSError, IOError):
            font_bold_large = ImageFont.load_default()
            font_bold = ImageFont.load_default()
            font_bold_small = ImageFont.load_default()
            font_small = ImageFont.load_default()
            font_tiny = ImageFont.load_default()

    def hex_to_rgb(hex_color):
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    def get_cell_bg(cell):
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
            rgb_str = cell.fill.start_color.rgb
            if isinstance(rgb_str, str) and len(rgb_str) >= 6:
                # Remove alpha if present
                if len(rgb_str) == 8:
                    rgb_str = rgb_str[2:]
                return hex_to_rgb(rgb_str)
        return (26, 26, 26)

    def get_cell_fg(cell):
        if cell.font and cell.font.color and cell.font.color.rgb:
            rgb_str = cell.font.color.rgb
            if isinstance(rgb_str, str) and len(rgb_str) >= 6:
                if len(rgb_str) == 8:
                    rgb_str = rgb_str[2:]
                return hex_to_rgb(rgb_str)
        return (255, 255, 255)

    # ─── Detect merged cells ────────────────────────────
    merged_ranges = list(ws.merged_cells.ranges)

    def is_merged(row, col):
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return mr
        return None

    # ─── Draw cells ─────────────────────────────────────
    y = padding
    for row_idx in range(1, ws.max_row + 1):
        rh = row_heights.get(row_idx, 30)
        x = padding

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cw = col_width_px

            # Skip non-origin merged cells
            mr = is_merged(row_idx, col_idx)
            if mr and (row_idx != mr.min_row or col_idx != mr.min_col):
                x += cw
                continue

            # For merged origin cells, compute full width/height
            actual_w = cw
            actual_h = rh
            if mr:
                actual_w = (mr.max_col - mr.min_col + 1) * cw
                actual_h = sum(row_heights.get(r, 30) for r in range(mr.min_row, mr.max_row + 1))

            # Draw background
            bg = get_cell_bg(cell)
            draw.rectangle([x, y, x + actual_w - 1, y + actual_h - 1], fill=bg)

            # Draw border
            border_color = (80, 80, 80)
            draw.rectangle([x, y, x + actual_w - 1, y + actual_h - 1], outline=border_color)

            # Draw text
            if cell.value:
                fg = get_cell_fg(cell)
                text = str(cell.value)

                # Choose font
                if row_idx == 1:
                    font = font_bold_large
                elif row_idx == 2:
                    font = font_bold
                elif cell.font and cell.font.bold:
                    if cell.font.size and cell.font.size <= 9:
                        font = font_bold_small
                    else:
                        font = font_bold_small
                elif cell.font and cell.font.size and cell.font.size <= 8:
                    font = font_tiny
                else:
                    font = font_small

                # Handle multi-line text
                lines = text.split('\n')
                total_text_h = len(lines) * (font.size + 4) if hasattr(font, 'size') else len(lines) * 14
                text_y = y + (actual_h - total_text_h) // 2

                for line in lines:
                    bbox = draw.textbbox((0, 0), line, font=font)
                    tw = bbox[2] - bbox[0]
                    text_x = x + (actual_w - tw) // 2
                    draw.text((text_x, text_y), line, fill=fg, font=font)
                    text_y += (font.size + 4) if hasattr(font, 'size') else 14

            x += cw
        y += rh

    img.save(output_image_path, 'PNG')
    print(f"✅ Created Image: {output_image_path}")
    return output_image_path


# ─── Main Execution ─────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("TKO DANCE STUDIO - Time Table Generator")
    print("=" * 60)

    # 1. Create Yatsuka Excel
    yatsuka_excel = create_timetable_excel(
        yatsuka_schedule,
        "谷塚スタジオ（YATSUKA STUDIO）",
        "timetable_yatsuka.xlsx",
        studio_color=RED,
        note_text="※ girl's HIP HOP は月2回の不定期開催です。開催日程の詳細はメールか電話でお問い合わせ下さい。"
    )

    # 2. Create Kitakoshigaya Excel
    koshigaya_excel = create_timetable_excel(
        kitakoshigaya_schedule,
        "北越谷スタジオ（KITAKOSHIGAYA STUDIO）",
        "timetable_kitakoshigaya.xlsx",
        studio_color=NAVY,
        note_text=None
    )

    # 3. Convert to images
    yatsuka_img = excel_to_image(
        yatsuka_excel,
        os.path.join(IMAGES_DIR, "schedule_yatsuka.png")
    )

    koshigaya_img = excel_to_image(
        koshigaya_excel,
        os.path.join(IMAGES_DIR, "schedule_kitakoshigaya.png")
    )

    print("\n" + "=" * 60)
    print("All files generated successfully!")
    print(f"  Excel files: {EXCEL_DIR}/")
    print(f"  Images:      {IMAGES_DIR}/")
    print("=" * 60)
